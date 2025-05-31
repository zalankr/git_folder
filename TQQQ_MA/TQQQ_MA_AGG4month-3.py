import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import load_workbook

if not os.path.exists('C:/Users/ilpus/PythonProjects/git_folder/TQQQ_MA/TQQQ_Results.xlsx'):
    with pd.ExcelWriter('C:/Users/ilpus/PythonProjects/git_folder/TQQQ_MA/TQQQ_Results.xlsx', engine='openpyxl') as writer:
        # 빈 데이터프레임으로 시트 생성
        pd.DataFrame().to_excel(writer, sheet_name='sheet1', index=False)

# 데이터 로드
# AGG 데이터
AGG = yf.download('AGG', start='2010-02-09', auto_adjust=True, interval='1mo', progress=True, multi_level_index=False)
AGG.drop(['Open','High','Low','Volume'], axis=1, inplace=True)

# Average
AGG.loc[:,'AGG_MA'] = AGG.loc[:,'Close'].rolling(window=4).mean()
AGG = AGG[['Close', 'AGG_MA']].rename(columns={'Close': 'AGG'})

AGG.loc[:,'Regime'] = AGG.loc[:,'AGG'].shift(1) >= AGG.loc[:,'AGG_MA'].shift(1)
AGG = AGG.resample('1d').ffill()
AGG['Regime'] = AGG['Regime'].ffill()
AGG.index = pd.to_datetime(AGG.index)
AGG['Regime'] = AGG['Regime'].astype(bool)
AGG = AGG.dropna(subset=['AGG_MA'])

# TQQQ 데이터
TQQQ = yf.download('TQQQ', start='2010-02-09', auto_adjust=True, interval='1d', progress=True, multi_level_index=False)
TQQQ.drop(['Open','High','Low','Volume'], axis=1, inplace=True)

# 반복 백테스트
for MA in range(5, 256, 5):
    # TQQQ의 이동평균 계산
    TQ = TQQQ.copy()
    AG = AGG.copy()
    TQ.loc[:,'TQ_MA'] = TQ.loc[:,'Close'].rolling(window=MA).mean()
    TQ = TQ[['Close', 'TQ_MA']].rename(columns={'Close': 'TQ'})
    TQ.loc[:,'Long'] = TQ.loc[:,'TQ'] >= TQ.loc[:,'TQ_MA']

    # 데이터 병합
    df = TQ.join(AG, how='left')
    df = df.sort_index()
    df.index = pd.to_datetime(df.index)
    df = df.dropna(subset=['TQ_MA', 'AGG_MA'])

    # 포지션
    df['Position'] = np.where((df['Long'] == 1) & (df['Regime'] == 1), 1, 0)
    df['MAPosition'] = np.where((df['Long'] == 1), 1, 0)
    df['Position'] = df['Position'].shift(1).fillna(0)
    df['MAPosition'] = df['MAPosition'].shift(1).fillna(0)

    # 데이터 정리
    df = df.drop(index=df.index[:225])
    df['daily_return'] = df['TQ'].pct_change().fillna(0)
    df['MA_return'] = df['daily_return'] * df['MAPosition']
    df['strategy_return'] = df['daily_return'] * df['Position']

    # 수수료
    fee = 0.0009
    df['MA_return'] -= df['MAPosition'].diff().abs().fillna(0) * fee
    df['strategy_return'] -= df['Position'].diff().abs().fillna(0) * fee

    # 누적 수익률
    df['cum_return'] = (1 + df['strategy_return']).cumprod()
    df['cum_MA_return'] = (1 + df['MA_return']).cumprod()
    df['cum_market'] = (1 + df['daily_return']).cumprod()

    # 거래 횟수
    df['signal'] = 0
    df.loc[df['Position'] == 1, 'signal'] = 1
    df.loc[df['Position'] == 0, 'signal'] = -1
    df['trade'] = df['signal'].diff().abs().fillna(0)
    df['trade'] = df['trade'].replace(2, 1)
    df['trade'] = df['trade'].astype(int)

    df.drop(['TQ_MA','AGG', 'AGG_MA', 'Regime'], axis=1, inplace=True)

    # 연도 계산
    start_date = df.index[0]
    end_date = df.index[-1]
    n_years = (end_date - start_date).days / 365.25

    # 통계 지표
    Strategy_results = []
    MA_results = []
    BH_results = []

    cagr = df['cum_return'].iloc[-1] ** (1 / n_years) - 1
    mdd = (df['cum_return'] / df['cum_return'].cummax() - 1).min()
    num_trades = df['trade'].sum()
    mean_return = df['strategy_return'].mean()
    std_return = df['strategy_return'].std()
    neg_std = df.loc[df['strategy_return'] < 0, 'strategy_return'].std()
    sharpe_ratio = (mean_return / std_return) * np.sqrt(252) if std_return != 0 else np.nan
    sortino_ratio = (mean_return / neg_std) * np.sqrt(252) if neg_std != 0 else np.nan

    buy_and_hold_return = df['cum_market'].iloc[-1] - 1
    buy_and_hold_cagr = buy_and_hold_return ** (1 / n_years) - 1
    buy_and_hold_mdd = (df['cum_market'] / df['cum_market'].cummax() - 1).min()
    BH_mean_return = df['daily_return'].mean()
    BH_std_return = df['daily_return'].std()
    neg_BH_std = df.loc[df['daily_return'] < 0, 'daily_return'].std()
    BH_sharpe_ratio = (BH_mean_return / BH_std_return) * np.sqrt(252) if BH_std_return != 0 else np.nan
    BH_sortino_ratio = (BH_mean_return / neg_BH_std) * np.sqrt(252) if neg_BH_std != 0 else np.nan

    MA_cagr = df['cum_MA_return'].iloc[-1] ** (1 / n_years) - 1
    MA_mdd = (df['cum_MA_return'] / df['cum_MA_return'].cummax() - 1).min()
    MA_mean_return = df['MA_return'].mean()
    MA_std_return = df['MA_return'].std()
    neg_MA_std = df.loc[df['MA_return'] < 0, 'MA_return'].std()
    MA_sharpe_ratio = (MA_mean_return / MA_std_return) * np.sqrt(252) if MA_std_return != 0 else np.nan
    MA_sortino_ratio = (MA_mean_return / neg_MA_std) * np.sqrt(252) if neg_MA_std != 0 else np.nan

    # 결과 저장
    Strategy_results.append({
        'Model': 'TQQQ MA+AGG',
        'MA': MA,
        'CAGR': cagr,
        'MDD': mdd,
        'Sharpe': sharpe_ratio,
        'Sortino': sortino_ratio,
        'Trades': int(num_trades)
    })

    MA_results.append({
        'Model': 'TQQQ MA',
        'MA': MA,
        'MA CAGR': MA_cagr,
        'MA MDD': MA_mdd,
        'MA Sharpe': MA_sharpe_ratio,
        'MA Sortino': MA_sortino_ratio,
    })

    BH_results.append({
        'Model': 'TQQQ BH',
        'CAGR': buy_and_hold_cagr,
        'MDD': buy_and_hold_mdd,
        'Sharpe': BH_sharpe_ratio,
        'Sortino': BH_sortino_ratio,
    })

    # 결과 정리
    Strategy_results = pd.DataFrame(Strategy_results)
    Strategy_results = Strategy_results.sort_values(by='CAGR', ascending=False).reset_index(drop=True)
    print(Strategy_results)

    MA_results = pd.DataFrame(MA_results)
    MA_results = MA_results.sort_values(by='MA CAGR', ascending=False).reset_index(drop=True)
    print(MA_results)

    BH_results = pd.DataFrame(BH_results)
    BH_results = BH_results.sort_values(by='CAGR', ascending=False).reset_index(drop=True)
    print(BH_results)

    results = pd.concat([Strategy_results, MA_results], axis=1)

    # 💾 기존 엑셀 파일 아래에 이어서 저장
    file_path = 'C:/Users/ilpus/PythonProjects/git_folder/TQQQ_MA/TQQQ_Results.xlsx'

    def get_start_row(sheet_name, file_path):
        if not os.path.exists(file_path):
            return 0
        wb = load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            return wb[sheet_name].max_row
        else:
            return 0
        
    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
        strategy_start_row = get_start_row('Strategy', file_path)
        bh_start_row = get_start_row('BH', file_path)

        results.to_excel(writer, sheet_name='Strategy', startrow=strategy_start_row, index=False, header=strategy_start_row == 0)
        BH_results.to_excel(writer, sheet_name='BH', startrow=bh_start_row, index=False, header=bh_start_row == 0)
